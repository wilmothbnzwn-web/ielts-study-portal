#!/usr/bin/env python3
"""
IELTS Study Portal — Desktop Vocabulary Import Script

Reads vocabulary from ~/Desktop/IELTS_Organized_Library and imports
new words into data/vocabulary.json. Re-runnable — skips duplicates.

Usage:
  python3 scripts/import-vocabulary-from-library.py

Sources:
  1. 雅思词汇EXCEL版-顺序版.xls (全部9400 sheet) — ~9400 words with phonetics
  2. 听力核心场景词汇.xlsx — ~655 listening scene words with categories
  3. 听力179词同义替换.xlsx — 179 listening test word-synonym pairs
  4. 雅思阅读-538考点词表格整理.xlsx — 538 reading test word-synonym pairs
  5. 雅思听力地图题词汇汇总.docx — 28 map navigation phrases
"""

import json
import re
import sys
import os

# ── Paths ──────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
VOCAB_PATH = os.path.join(PROJECT_DIR, 'data', 'vocabulary.json')
DESKTOP_LIB = os.path.expanduser('~/Desktop/IELTS_Organized_Library')

# ── Scene → Theme mapping ──────────────────────────────────────────
SCENE_THEME_MAP = {
    'Rental & Accommodation 租赁和住宿': 'Campus Life — Accommodation',
    'Public Facilities 公共设施': 'Campus Life — Facilities',
    'Registration & Enrolment 注册和入学': 'Campus Life — Enrolment',
    'Course Selection 课程选择': 'Campus Life — Courses',
    'Internship/Academic Feedback 实习/学业反馈': 'Campus Life — Academic Support',
    'Consultation 咨询': 'Campus Life — Services',
    'Appointment & Booking 预约和预定': 'Campus Life — Services',
    'Biology 生物': 'Academic Subject — Biology',
    'Mathematics, Chemistry & Physics 数学, 化学和物理': 'Academic Subject — Science',
    'Psychology & Social Sciences 心理学和社会科学': 'Academic Subject — Social Sciences',
    'History, Archaeology & Humanities 历史, 考古和人文': 'Academic Subject — Humanities',
    'Geography 地理': 'Academic Subject — Geography',
    'Arts & Music 艺术与音乐': 'Academic Subject — Arts',
    'Education & Languages 教育和语言': 'Academic Subject — Education',
    'Energy & Environment 能源和环境': 'Academic Subject — Environment',
    'Architecture & Design 建筑与设计': 'Academic Subject — Architecture',
    'Engineering 工程': 'Academic Subject — Engineering',
    'Management & Business 管理和商业': 'Academic Subject — Business',
    'Culture, Literature & Media 文化, 文学和媒体': 'Academic Subject — Humanities',
    'Technology 技术': 'Academic Subject — Technology',
    'Jobs & Employment 工作和就业': 'Work & Employment',
    'Survey & Research 调查与研究': 'Academic Research',
    'Shopping 购物': 'Daily Life — Shopping',
    'Food & Restaurant 食物和餐厅': 'Daily Life — Food',
    'Traffic 交通': 'Daily Life — Transport',
    'Tourism 旅游': 'Travel & Tourism',
    'Sports & Fitness 运动与健康': 'Health & Fitness',
    'Health & Medicine 健康与医疗': 'Health & Medicine',
    'Entertainment 娱乐': 'Leisure & Entertainment',
    'Activity & Festival 活动和节日': 'Leisure & Activities',
}

def map_scene(scene_name):
    for key, theme in SCENE_THEME_MAP.items():
        if key.startswith(scene_name) or scene_name.startswith(key):
            return theme
    return f'IELTS Listening: {scene_name}'

def parse_pos_and_meaning(raw):
    """Parse 'vt. 抛弃;放弃' → (pos, chinese)"""
    if not raw:
        return '', ''
    raw = raw.strip()
    m = re.match(
        r'^(n\.|v\.|vt\.|vi\.|adj\.|adv\.|prep\.|conj\.|pron\.|int\.|art\.|num\.|aux\.|a\.|ad\.|abbr\.)\s*(.*)',
        raw, re.IGNORECASE
    )
    if m:
        pos = m.group(1).strip()
        meaning = m.group(2).strip()
        pos_map = {
            'a.': 'adj.', 'ad.': 'adv.', 'v.': 'v.', 'vt.': 'vt.', 'vi.': 'vi.',
            'n.': 'n.', 'adj.': 'adj.', 'adv.': 'adv.', 'prep.': 'prep.',
            'conj.': 'conj.', 'pron.': 'pron.', 'int.': 'int.', 'abbr.': 'abbr.'
        }
        pos = pos_map.get(pos.lower(), pos)
        return pos, meaning
    return '', raw


class VocabImporter:
    def __init__(self, vocab_path):
        with open(vocab_path, 'r', encoding='utf-8') as f:
            self.data = json.load(f)
        self.existing_words = set(d['word'].lower().strip() for d in self.data)
        self.max_id = max(d['id'] for d in self.data) if self.data else 0
        self.stats = {
            'files_read': [],
            'duplicates_skipped': 0,
            'invalid_skipped': 0,
            'new_imported': 0,
            'per_source': {},
        }

    def add_entry(self, word, chinese='', pos='', synonyms=None, definition='',
                  example='', theme='', difficulty=3, source='IELTS Desktop Library'):
        """Add candidate if not already present (case-insensitive)"""
        if not word or len(word.strip()) < 2:
            self.stats['invalid_skipped'] += 1
            return
        word = word.strip()
        key = word.lower().strip()

        if key in self.existing_words:
            self.stats['duplicates_skipped'] += 1
            return

        self.max_id += 1
        entry = {
            'id': self.max_id,
            'word': word,
            'pos': pos,
            'chinese': chinese or '',
            'synonyms': synonyms if synonyms else [],
            'definition': definition or '',
            'example': example or '',
            'theme': theme or '',
            'difficulty': difficulty,
            'source': source,
        }
        self.data.append(entry)
        self.existing_words.add(key)
        self.stats['new_imported'] += 1
        self.stats['per_source'][source] = self.stats['per_source'].get(source, 0) + 1

    def import_9400_xls(self):
        """Import from 雅思词汇EXCEL版-顺序版.xls (全部9400 sheet)"""
        import xlrd
        xls_path = os.path.join(
            DESKTOP_LIB,
            '05_高频词汇与替换语料库_Vocab_and_Synonyms',
            '02_词汇书与词表/雅思英语单词表/雅思词汇EXCEL版-顺序版.xls'
        )
        if not os.path.exists(xls_path):
            print(f"  [SKIP] File not found: {xls_path}")
            return
        self.stats['files_read'].append('雅思词汇EXCEL版-顺序版.xls (全部9400)')
        wb = xlrd.open_workbook(xls_path)
        ws = wb.sheet_by_name('全部9400')
        count = 0
        for r in range(1, ws.nrows):
            row = ws.row_values(r)
            if len(row) < 4:
                continue
            word = str(row[1]).strip() if row[1] else ''
            raw_chinese = str(row[3]).strip() if row[3] else ''
            if not word or word.startswith('('):
                continue
            pos, chinese = parse_pos_and_meaning(raw_chinese)
            self.add_entry(word=word, chinese=chinese, pos=pos,
                           source='IELTS 9400 Core Vocabulary')
            count += 1
        print(f"  9400 XLS: parsed {count} rows, +{self.stats['per_source'].get('IELTS 9400 Core Vocabulary', 0)} new")

    def import_scene_xlsx(self):
        """Import from 听力核心场景词汇.xlsx"""
        import openpyxl
        xlsx_path = os.path.join(
            DESKTOP_LIB,
            '05_高频词汇与替换语料库_Vocab_and_Synonyms',
            '01_同义词替换合集/听力核心场景词汇.xlsx'
        )
        if not os.path.exists(xlsx_path):
            print(f"  [SKIP] File not found: {xlsx_path}")
            return
        self.stats['files_read'].append('听力核心场景词汇.xlsx')
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4 or not row[2]:
                continue
            scene = str(row[1]).strip() if row[1] else ''
            word = str(row[2]).strip()
            chinese = str(row[3]).strip() if row[3] else ''
            theme = map_scene(scene)
            self.add_entry(word=word, chinese=chinese, theme=theme,
                           source='IELTS听力场景分类词汇', difficulty=2)
            count += 1
        print(f"  Scene XLSX: parsed {count} rows, +{self.stats['per_source'].get('IELTS听力场景分类词汇', 0)} new")

    def import_179_xlsx(self):
        """Import from 听力179词同义替换.xlsx"""
        import openpyxl
        xlsx_path = os.path.join(
            DESKTOP_LIB,
            '05_高频词汇与替换语料库_Vocab_and_Synonyms',
            '03_考点词系列_538_179/179听力考点词/听力179词同义替换.xlsx'
        )
        if not os.path.exists(xlsx_path):
            print(f"  [SKIP] File not found: {xlsx_path}")
            return
        self.stats['files_read'].append('听力179词同义替换.xlsx')
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[1]:
                continue
            word = str(row[1]).strip()
            if word.lower() in ['序号', 't', '']:
                continue
            meaning = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            synonyms_str = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            synonyms = [s.strip() for s in synonyms_str.replace('，', ',').split(',') if s.strip()] if synonyms_str else []
            self.add_entry(word=word, chinese=meaning, synonyms=synonyms,
                           source='听力179考点同义替换')
            count += 1
        print(f"  179 XLSX: parsed {count} rows, +{self.stats['per_source'].get('听力179考点同义替换', 0)} new")

    def import_538_xlsx(self):
        """Import from 雅思阅读-538考点词表格整理.xlsx"""
        import openpyxl
        xlsx_path = os.path.join(
            DESKTOP_LIB,
            '05_高频词汇与替换语料库_Vocab_and_Synonyms',
            '03_考点词系列_538_179/538阅读考点词/雅思阅读-538考点词表格整理.xlsx'
        )
        if not os.path.exists(xlsx_path):
            print(f"  [SKIP] File not found: {xlsx_path}")
            return
        self.stats['files_read'].append('雅思阅读-538考点词表格整理.xlsx')
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[1]:
                continue
            word = str(row[1]).strip()
            if word.lower() in ['1类', 'part 2', 'part 3', '']:
                continue
            definition_raw = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            synonyms_str = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            pos, chinese = parse_pos_and_meaning(definition_raw)
            synonyms = [s.strip() for s in synonyms_str.replace('，', ',').replace(';', ',').split(',') if s.strip()] if synonyms_str else []
            self.add_entry(word=word, chinese=chinese, pos=pos, synonyms=synonyms,
                           source='阅读538考点词真经')
            count += 1
        print(f"  538 XLSX: parsed {count} rows, +{self.stats['per_source'].get('阅读538考点词真经', 0)} new")

    def import_map_docx(self):
        """Import from 雅思听力地图题词汇汇总.docx"""
        from docx import Document
        docx_path = os.path.join(
            DESKTOP_LIB,
            '05_高频词汇与替换语料库_Vocab_and_Synonyms',
            '01_同义词替换合集/雅思听力地图题词汇汇总.docx'
        )
        if not os.path.exists(docx_path):
            print(f"  [SKIP] File not found: {docx_path}")
            return
        self.stats['files_read'].append('雅思听力地图题词汇汇总.docx')
        doc = Document(docx_path)
        phrases = []
        current_category = ''
        for p in doc.paragraphs:
            text = p.text.strip()
            if not text:
                continue
            if re.match(r'^[一二三四五六七八九十]、', text):
                current_category = text
                continue
            m = re.match(r'^(\d+)[.,、]\s*(.+?)\s{2,}(.+)$', text)
            if m:
                phrase_en = m.group(2).strip()
                phrase_cn = m.group(3).strip()
                phrases.append((phrase_en, phrase_cn))
            else:
                m2 = re.match(r'^(\d+)[.,、]\s*(.+)$', text)
                if m2:
                    content = m2.group(2).strip()
                    split_idx = None
                    for i, ch in enumerate(content):
                        if '一' <= ch <= '鿿' or '㐀' <= ch <= '䶿':
                            split_idx = i
                            break
                    if split_idx is not None:
                        phrase_en = content[:split_idx].strip().rstrip('.,; ')
                        phrase_cn = content[split_idx:].strip()
                        phrases.append((phrase_en, phrase_cn))
        count = 0
        for phrase_en, phrase_cn in phrases:
            if phrase_en and len(phrase_en) >= 2:
                self.add_entry(word=phrase_en, chinese=phrase_cn, pos='phrase',
                               theme='IELTS Listening — Map Navigation',
                               source='IELTS听力地图题词汇', difficulty=2)
                count += 1
        print(f"  Map DOCX: parsed {count} phrases, +{self.stats['per_source'].get('IELTS听力地图题词汇', 0)} new")

    def save(self):
        """Write updated data back to vocabulary.json"""
        # Remove any accidental header-like entries
        self.data = [d for d in self.data
                     if d['word'].lower().strip() not in ['part 2', 'part 3', '1类', '2类']
                     and len(d['word'].strip()) >= 2]

        # Re-assign sequential IDs
        for i, d in enumerate(self.data):
            d['id'] = i + 1

        with open(VOCAB_PATH, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    def print_report(self):
        print(f"\n{'='*60}")
        print(f"  Import Report")
        print(f"{'='*60}")
        print(f"  Files read: {len(self.stats['files_read'])}")
        for f in self.stats['files_read']:
            print(f"    - {f}")
        print(f"  Total parsed from desktop: (see per-source counts)")
        print(f"  Existing before import: {len(self.data) - self.stats['new_imported']}")
        print(f"  New imported: {self.stats['new_imported']}")
        print(f"  Duplicates skipped: {self.stats['duplicates_skipped']}")
        print(f"  Invalid skipped: {self.stats['invalid_skipped']}")
        print(f"  Final total: {len(self.data)}")
        print(f"\n  Per-source breakdown:")
        for src, cnt in sorted(self.stats['per_source'].items(), key=lambda x: -x[1]):
            print(f"    {src}: +{cnt}")
        print(f"{'='*60}")


def main():
    print("IELTS Study Portal — Desktop Vocabulary Import")
    print(f"Desktop library: {DESKTOP_LIB}")
    print(f"Target: {VOCAB_PATH}\n")

    if not os.path.exists(DESKTOP_LIB):
        print("ERROR: Desktop library not found at", DESKTOP_LIB)
        print("Please ensure ~/Desktop/IELTS_Organized_Library exists.")
        sys.exit(1)

    importer = VocabImporter(VOCAB_PATH)
    initial_count = len(importer.data)

    print("Importing...")
    importer.import_9400_xls()
    importer.import_scene_xlsx()
    importer.import_179_xlsx()
    importer.import_538_xlsx()
    importer.import_map_docx()

    importer.save()
    importer.print_report()

    if importer.stats['new_imported'] > 0:
        print(f"\nDone. Added {importer.stats['new_imported']} new words.")
        print("Restart the dev server to pick up changes.")
    else:
        print("\nNo new words to import — everything is already in vocabulary.json.")


if __name__ == '__main__':
    main()
